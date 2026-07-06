"""
app.py
------
A small internal web app that reports on developer activity in Jira.

Run:
    pip install -r requirements.txt
    export JIRA_BASE_URL=https://lifedata.atlassian.net
    export JIRA_EMAIL=you@lifedatacorp.com
    export JIRA_API_TOKEN=*****
    python app.py
    # open http://localhost:5000

Pages:
    /                      team overview + per-developer summary table
    /developer/<name>      one developer's completed + in-progress detail
    /report.xlsx           download the whole thing as an Excel workbook
    /api/report.json       raw JSON (handy for a future dashboard/automation)

The numbers come from jira_client.build_report(), which reads the Jira changelog
to compute true In-Progress -> Done cycle time and time-in-current-status.
Results are cached for a few minutes so page loads are fast and gentle on the API.
"""

from __future__ import annotations

import csv
import io
import time

from flask import Flask, Response, abort, jsonify, render_template_string, request

import config as cfg
import jira_client as jc
import reports as R
import reports_web

import dev_reports_web

app = Flask(__name__)
# Eight executive reports (daily movement, sprint, dev/QA productivity, status
# duration, release readiness, executive dashboard, individual activity) live here.
app.register_blueprint(reports_web.bp)
# The 18 developer-discipline reports (Jira Developer Reports spec).
app.register_blueprint(dev_reports_web.devbp)
# v3 screens: Settings, My Day, Attention, QA, Flow, Quality, Planning, Investigator.
import screens_web
app.register_blueprint(screens_web.v3)

# ---- tiny in-memory cache so we don't hammer the Jira API on every refresh ----
_CACHE: dict = {"data": None, "ts": 0.0}
_CACHE_TTL = 300  # seconds


def get_reports(force: bool = False):
    if force or not _CACHE["data"] or (time.time() - _CACHE["ts"] > _CACHE_TTL):
        _CACHE["data"] = jc.build_report(fetch_changelogs=True)
        _CACHE["ts"] = time.time()
    return _CACHE["data"]


def fmt(v, suffix="d"):
    return f"{v}{suffix}" if v is not None else "—"


def hdur(days):
    """Human-friendly duration from a float number of days: '3d 4h', '18h', '25m'."""
    if days is None:
        return "—"
    total_h = days * 24
    if total_h < 1:
        return f"{max(round(total_h * 60), 1)}m"
    if total_h < 24:
        return f"{round(total_h)}h"
    d = int(total_h // 24)
    h = int(round(total_h - d * 24))
    return f"{d}d {h}h" if h else f"{d}d"


def agecls(days, warn=10, bad=20):
    """CSS class for aging: '' (fresh), 'warn' (>=10d), 'bad' (>=20d)."""
    if days is None:
        return ""
    if days >= bad:
        return "bad"
    if days >= warn:
        return "warn"
    return ""


# ---------------------------------------------------------------------------
# Templates (kept inline to keep this a small, copy-pasteable project)
# ---------------------------------------------------------------------------

BASE_CSS = """
<style>
  body { font-family: -apple-system, Segoe UI, Roboto, sans-serif; margin: 0; color: #172b4d; background: #f4f5f7; }
  header { background: #0052cc; color: #fff; padding: 18px 28px; }
  header h1 { margin: 0; font-size: 20px; }
  header .sub { opacity: .85; font-size: 13px; margin-top: 4px; }
  .wrap { max-width: 1100px; margin: 24px auto; padding: 0 20px; }
  .cards { display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 24px; }
  .card { background: #fff; border-radius: 8px; padding: 16px 20px; box-shadow: 0 1px 3px rgba(9,30,66,.12); flex: 1; min-width: 160px; }
  .card .n { font-size: 28px; font-weight: 700; }
  .card .l { color: #6b778c; font-size: 13px; margin-top: 2px; }
  table { width: 100%; border-collapse: collapse; background: #fff; border-radius: 8px; overflow: hidden; box-shadow: 0 1px 3px rgba(9,30,66,.12); }
  th, td { text-align: left; padding: 10px 14px; border-bottom: 1px solid #ebecf0; font-size: 14px; }
  th { background: #fafbfc; color: #6b778c; font-weight: 600; }
  tr:hover td { background: #f7f8fa; }
  a { color: #0052cc; text-decoration: none; }
  a:hover { text-decoration: underline; }
  .pill { display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 12px; background: #dfe1e6; }
  .pill.warn { background:#fff7e6; color:#974f00; }
  .pill.bad { background:#ffebe6; color:#bf2600; }
  .pill.ok { background:#e3fcef; color:#006644; }
  .badge-rework { display:inline-block;background:#ffebe6;color:#bf2600;border-radius:10px;padding:2px 8px;font-size:11px;font-weight:500;margin-left:6px; }
  .search { width:100%;max-width:340px;padding:8px 12px;border:1px solid #dfe1e6;border-radius:8px;font-size:14px;margin-bottom:14px; }
  .search:focus { outline:none;border-color:#4c9aff;box-shadow:0 0 0 2px rgba(76,154,255,.25); }
  th[title], .help { border-bottom:1px dotted #b3bac5;cursor:help; }
  details.txns summary { cursor:pointer;font-size:13px;color:#0052cc;padding:4px 0;user-select:none; }
  details.txns[open] summary { margin-bottom:6px; }
  .attention { display:flex;gap:10px;flex-wrap:wrap;margin-bottom:18px; }
  .attention .a { background:#fff;border:1px solid #e3e6ea;border-radius:8px;padding:8px 14px;font-size:13px;color:#5e6c84; }
  .attention .a b { font-size:16px;color:#172b4d;margin-right:5px; }
  .attention .a.hot { background:#ffebe6;border-color:#ffbdad; } .attention .a.hot b { color:#bf2600; }
  .attention .a.warm { background:#fff7e6;border-color:#ffe2b3; } .attention .a.warm b { color:#974f00; }
  .journey { display:flex;height:24px;border-radius:6px;overflow:hidden;font-size:11px;color:#fff;font-weight:500;margin:4px 0 6px; }
  .journey .seg { display:flex;align-items:center;justify-content:center;white-space:nowrap;overflow:hidden;min-width:0;text-shadow:0 1px 1px rgba(0,0,0,.35); }
  .legend { display:flex;gap:14px;flex-wrap:wrap;font-size:12px;color:#5e6c84;margin-bottom:8px; }
  .legend .sw { display:inline-block;width:10px;height:10px;border-radius:2px;margin-right:4px;vertical-align:-1px; }
  table.sortable th { cursor:pointer;user-select:none; }
  table.sortable th .ind { color:#0052cc;font-size:10px; }
  .controls { display:flex;align-items:center;gap:8px;flex-wrap:wrap;margin-bottom:12px; }
  .controls .ctl-label { font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.04em;color:#5e6c84;margin-right:2px; }
  .chipbtn { background:#fff;border:1px solid #dfe1e6;border-radius:16px;padding:5px 12px;font-size:13px;color:#42526e;cursor:pointer; }
  .chipbtn:hover { border-color:#9fb3d1; }
  .chipbtn.active { background:#e6effd;border-color:#4c9aff;color:#0747a6;font-weight:500; }
  .group-header { font-size:13px;font-weight:500;color:#172b4d;margin:16px 0 8px;padding-bottom:4px;border-bottom:2px solid #dfe1e6; }
  .bench { margin:5px 0 0; }
  .warn { background: #ffebe6; color: #bf2600; }
  .muted { color: #6b778c; font-size: 13px; }
  h2 { font-size: 16px; margin: 28px 0 10px; }
  .toolbar { margin-bottom: 16px; }
  .btn { background:#0052cc;color:#fff;padding:8px 14px;border-radius:6px;font-size:13px;border:none;cursor:pointer;display:inline-block;text-decoration:none; }
  .btn:hover { background:#0747a6;text-decoration:none; }
  /* ---- filter bar ---- */
  .filterbar { background:#fff;border:1px solid #e3e6ea;border-radius:10px;margin-bottom:22px;box-shadow:0 1px 3px rgba(9,30,66,.12);overflow:hidden; }
  .filterbar-head { display:flex;align-items:center;justify-content:space-between;padding:11px 18px;border-bottom:1px solid #ebecf0;background:#fafbfc; }
  .filterbar-head h3 { margin:0;font-size:12px;font-weight:700;letter-spacing:.05em;text-transform:uppercase;color:#5e6c84; }
  .filter-body { padding:6px 18px; }
  .filter-group { display:flex;align-items:flex-start;gap:16px;padding:11px 0; }
  .filter-group + .filter-group { border-top:1px dashed #ebecf0; }
  .filter-key { min-width:110px;font-size:13px;font-weight:600;color:#172b4d;padding-top:6px; }
  .chips { display:flex;flex-wrap:wrap;gap:8px;flex:1; }
  .chip { display:inline-flex;align-items:center;gap:7px;padding:5px 12px;border:1px solid #dfe1e6;border-radius:18px;cursor:pointer;font-size:13px;color:#42526e;background:#fff;transition:background .12s,border-color .12s,color .12s;user-select:none; }
  .chip:hover { border-color:#9fb3d1; }
  .chip input { accent-color:#0052cc;width:14px;height:14px;margin:0;cursor:pointer; }
  .chip:has(input:checked) { background:#e6effd;border-color:#4c9aff;color:#0747a6;font-weight:600; }
  .num-input { width:66px;padding:6px 8px;border:1px solid #dfe1e6;border-radius:6px;font-size:13px; }
  .num-input:focus { outline:none;border-color:#4c9aff;box-shadow:0 0 0 2px rgba(76,154,255,.25); }
  .filter-actions { display:flex;align-items:center;gap:10px;padding:11px 18px;border-top:1px solid #ebecf0;background:#fafbfc; }
  .btn-ghost { background:#fff;color:#42526e;border:1px solid #dfe1e6;padding:7px 14px;border-radius:6px;font-size:13px;cursor:pointer;text-decoration:none;display:inline-block; }
  .btn-ghost:hover { background:#f4f5f7;border-color:#b3bac5;text-decoration:none; }
</style>
""" + reports_web.LOADING_OVERLAY

OVERVIEW_TMPL = BASE_CSS + """
<header>
  <h1>Developer Activity Report</h1>
  <div class="sub">{{ projects }} &middot; completed in last {{ window }} days &middot; generated {{ generated }}</div>
</header>
<div class="wrap">
  <div class="cards">
    <div class="card"><div class="n">{{ total_done }}</div><div class="l">Tickets completed</div></div>
    <div class="card"><div class="n">{{ total_wip }}</div><div class="l">In progress now</div></div>
    <div class="card"><div class="n">{{ team_cycle }}</div><div class="l">Median cycle time</div></div>
    <div class="card"><div class="n">{{ team_oldest }}</div><div class="l">Oldest WIP (in status)</div></div>
  </div>

  <div class="toolbar">
    <a class="btn" href="/exec">Executive dashboard &amp; reports →</a>
    <a class="btn" href="/report.xlsx" download>Download Excel</a>
  </div>

  <table>
    <tr>
      <th>Developer</th><th>Open assigned</th><th>Completed</th><th>Avg cycle</th><th>Median cycle</th>
      <th>In progress</th><th>Oldest WIP</th>
    </tr>
    {% for d in devs %}
    <tr>
      <td><a href="/developer/{{ d.name|urlencode }}">{{ d.name }}</a></td>
      <td>{{ d.open_count }}</td>
      <td>{{ d.throughput }}</td>
      <td>{{ fmt(d.avg_cycle) }}</td>
      <td>{{ fmt(d.median_cycle) }}</td>
      <td>{{ d.in_progress|length }}</td>
      <td>{{ fmt(d.oldest_in_progress) }}</td>
    </tr>
    {% endfor %}
  </table>
  <p class="muted">Cycle time = first entry into an In&nbsp;Progress status &rarr; Done, from the issue changelog.
  Oldest WIP = days the ticket has sat in its current status. Grouped by current assignee.</p>
</div>
"""

DEV_TMPL = BASE_CSS + """
<header>
  <h1>{{ d.name }}</h1>
  <div class="sub"><a href="/" style="color:#cfe0ff">&larr; All developers</a></div>
</header>
<div class="wrap">
  <a class="btn" href="/developer/{{ d.name|urlencode }}/history" style="display:inline-block;margin-bottom:18px">View full activity history &rarr;</a>
  <div class="attention">
    <span class="a {{ 'hot' if alerts.stuck else '' }}"><b>{{ alerts.stuck }}</b>stuck &ge;10d</span>
    <span class="a {{ 'hot' if alerts.reopened else '' }}"><b>{{ alerts.reopened }}</b>reopened</span>
    <span class="a {{ 'warm' if alerts.blocked else '' }}"><b>{{ alerts.blocked }}</b>blocked</span>
    <span class="a"><b>{{ hdur(alerts.oldest) if alerts.oldest else '0d' }}</b>oldest open ticket</span>
  </div>
  <form method="get" class="filterbar">
    <div class="filterbar-head">
      <h3>Filters</h3>
      <a class="btn-ghost" href="{{ request.path }}/report.csv?{{ request.query_string.decode() }}" download>⬇ Download CSV</a>
    </div>
    <div class="filter-body">
      <div class="filter-group">
        <div class="filter-key">Ticket type</div>
        <div class="chips">
          {% for t in avail_types %}
          <label class="chip"><input type="checkbox" name="type" value="{{ t }}" {% if t in sel_types %}checked{% endif %}>{{ t }}</label>
          {% else %}<span class="muted">none</span>{% endfor %}
        </div>
      </div>
      <div class="filter-group">
        <div class="filter-key">Status</div>
        <div class="chips">
          {% for s in avail_statuses %}
          <label class="chip"><input type="checkbox" name="status" value="{{ s }}" {% if s in sel_statuses %}checked{% endif %}>{{ s }}</label>
          {% else %}<span class="muted">none</span>{% endfor %}
        </div>
      </div>
      <div class="filter-group">
        <div class="filter-key">Min open age</div>
        <div class="chips" style="align-items:center">
          <input type="number" name="min_age" min="0" value="{{ min_age }}" class="num-input"> <span class="muted">days</span>
        </div>
      </div>
    </div>
    <div class="filter-actions">
      <button class="btn" type="submit">Apply filters</button>
      <a class="btn-ghost" href="{{ request.path }}">Clear</a>
    </div>
  </form>
  <div class="cards">
    <div class="card"><div class="n">{{ d.open_count }}</div><div class="l">Open assigned</div></div>
    <div class="card"><div class="n">{{ d.throughput }}</div><div class="l">Completed ({{ window }}d)</div></div>
    <div class="card"><div class="n">{{ hdur(d.median_cycle) }}</div><div class="l help" title="Median first-In-Progress to Done across completed tickets">Median cycle time</div></div>
    <div class="card"><div class="n">{{ d.in_progress|length }}</div><div class="l">In progress</div></div>
    <div class="card"><div class="n">{{ hdur(d.oldest_in_progress) }}</div><div class="l help" title="Longest any in-progress ticket has sat in its current status">Oldest WIP</div></div>
  </div>

  <input id="devSearch" class="search" type="search" placeholder="Search tickets by key or summary…" autocomplete="off">

  <h2>In progress</h2>
  <table class="searchable sortable">
    <tr><th>Key</th><th>Summary</th><th>Current status</th><th title="Time the ticket has sat in its current status (aging)">Days in status</th><th title="Total time in active stages — actual work time, excluding paused/blocked">Time in progress</th></tr>
    {% for t in d.in_progress %}
    <tr>
      <td><a href="{{ t.url }}" target="_blank">{{ t.key }}</a>{% if t.reopened %}<span class="badge-rework" title="Reopened/sent back {{ t.reopened }} time(s)">&#8617; {{ t.reopened }}&times;</span>{% endif %}</td>
      <td>{{ t.summary }}</td>
      <td>{{ t.status }}</td>
      <td data-sort="{{ t.days_in_status or 0 }}">{% set c = agecls(t.days_in_status) %}<span class="{{ 'pill ' + c if c else '' }}" title="{{ t.days_in_status }} days">{{ hdur(t.days_in_status) }}</span></td>
      <td data-sort="{{ t.active_days or 0 }}" title="{{ t.active_days }} days">{{ hdur(t.active_days) }}</td>
    </tr>
    {% else %}<tr><td colspan="5" class="muted">Nothing in progress.</td></tr>{% endfor %}
  </table>
  <p class="muted"><b>Days in status</b> = time in the current status (aging; <span class="pill warn">amber &ge;10d</span> <span class="pill bad">red &ge;20d</span>). <b>Time in progress</b> = total time in active/in-progress stages — actual work time, excluding paused/blocked.</p>

  <h2>Completed</h2>
  <table class="searchable sortable">
    <tr><th>Key</th><th>Summary</th><th>Type</th><th title="Created → resolved (calendar time)">Lead</th><th title="First In Progress → resolved (active cycle time)">Cycle</th></tr>
    {% for t in d.completed %}
    <tr>
      <td><a href="{{ t.url }}" target="_blank">{{ t.key }}</a>{% if t.reopened %}<span class="badge-rework" title="Reopened/sent back {{ t.reopened }} time(s)">&#8617; {{ t.reopened }}&times;</span>{% endif %}</td>
      <td>{{ t.summary }}</td>
      <td><span class="pill">{{ t.issue_type }}</span></td>
      <td data-sort="{{ t.lead_days or 0 }}" title="{{ t.lead_days }} days">{{ hdur(t.lead_days) }}</td>
      <td data-sort="{{ t.cycle_days or 0 }}" title="{{ t.cycle_days }} days">{{ hdur(t.cycle_days) }}</td>
    </tr>
    {% else %}<tr><td colspan="5" class="muted">None in window.</td></tr>{% endfor %}
  </table>

  <h2>Currently assigned <span class="muted">(all open tickets)</span></h2>
  <table class="searchable sortable">
    <tr><th>Key</th><th>Summary</th><th>Type</th><th>Status</th><th title="Calendar days since the ticket was created">Open age</th></tr>
    {% for t in d.assigned %}
    <tr>
      <td><a href="{{ t.url }}" target="_blank">{{ t.key }}</a></td>
      <td>{{ t.summary }}</td>
      <td><span class="pill">{{ t.issue_type }}</span></td>
      <td>{{ t.status }}</td>
      <td data-sort="{{ t.age_days or 0 }}">{% set c = agecls(t.age_days) %}<span class="{{ 'pill ' + c if c else '' }}" title="{{ t.age_days }} days">{{ hdur(t.age_days) }}</span></td>
    </tr>
    {% else %}<tr><td colspan="5" class="muted">No open tickets assigned.</td></tr>{% endfor %}
  </table>
</div>
<script>
(function(){
  var box=document.getElementById('devSearch');
  if(box){box.addEventListener('input',function(){
    var q=box.value.toLowerCase();
    document.querySelectorAll('table.searchable').forEach(function(tb){
      var rows=tb.querySelectorAll('tr');
      for(var i=1;i<rows.length;i++){
        var r=rows[i];
        if(r.querySelector('td[colspan]'))continue;
        r.style.display = r.textContent.toLowerCase().indexOf(q)>=0 ? '' : 'none';
      }
    });
  });}
  function cellVal(row,col){
    var c=row.children[col]; if(!c)return '';
    var ds=c.getAttribute('data-sort');
    if(ds!==null)return parseFloat(ds)||0;
    return c.textContent.trim().toLowerCase();
  }
  document.querySelectorAll('table.sortable').forEach(function(table){
    var header=table.querySelector('tr'); if(!header)return;
    Array.prototype.forEach.call(header.children,function(th,col){
      th.addEventListener('click',function(){
        var asc=th.getAttribute('data-dir')!=='asc';
        Array.prototype.forEach.call(header.children,function(h){
          h.removeAttribute('data-dir'); var i=h.querySelector('.ind'); if(i)i.remove();
        });
        th.setAttribute('data-dir',asc?'asc':'desc');
        var ind=document.createElement('span'); ind.className='ind';
        ind.textContent=asc?' ▲':' ▼'; th.appendChild(ind);
        var rows=Array.prototype.slice.call(table.querySelectorAll('tr')).filter(function(r){
          return r!==header && !r.querySelector('td[colspan]');
        });
        rows.sort(function(a,b){
          var av=cellVal(a,col), bv=cellVal(b,col);
          return av<bv?(asc?-1:1):(av>bv?(asc?1:-1):0);
        });
        rows.forEach(function(r){table.appendChild(r);});
      });
    });
  });
})();
</script>
"""


HIST_TMPL = BASE_CSS + """
<style>
 .hist-card{background:#fff;border:1px solid #e3e6ea;border-radius:10px;box-shadow:0 1px 3px rgba(9,30,66,.12);padding:14px 18px;margin-bottom:16px}
 .hist-head{display:flex;justify-content:space-between;align-items:flex-start;gap:12px;flex-wrap:wrap}
 .hist-key{font-weight:700;margin-right:8px}
 .hist-tags{display:flex;gap:6px;flex-shrink:0}
 .hist-meta{color:#5e6c84;font-size:13px;margin:6px 0 12px}
 .hist-grid{display:grid;grid-template-columns:minmax(0,1fr) minmax(0,1.5fr);gap:18px}
 .hist-label{font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.04em;color:#5e6c84;margin-bottom:6px}
 table.mini{width:100%;border-collapse:collapse;background:#fff;box-shadow:none}
 table.mini th{font-size:11px;text-transform:uppercase;letter-spacing:.03em;color:#6b778c;text-align:left;padding:6px 8px;border-bottom:1px solid #ebecf0;font-weight:600}
 table.mini td{padding:5px 8px;border-bottom:1px solid #f1f2f4;font-size:13px}
 @media (max-width:760px){.hist-grid{grid-template-columns:1fr}}
</style>
<header>
  <h1>Activity history — {{ d.person }}</h1>
  <div class="sub"><a href="/" style="color:#cfe0ff">&larr; Overview</a> &middot; <a href="/developer/{{ d.person|urlencode }}" style="color:#cfe0ff">Back to summary</a> &middot; complete per-ticket history &middot; lookback {{ days }} days</div>
</header>
<div class="wrap">
  <div class="cards">
    <div class="card"><div class="n">{{ d.ticket_count }}</div><div class="l">Tickets worked on</div></div>
    <div class="card"><div class="n">{{ hdur(d.active_days_total) }}</div><div class="l help" title="Total time across all of this person's tickets spent in active stages">Total active time</div></div>
  </div>
  {% if d.insight.top_stage %}
  <p class="muted" style="font-size:14px;margin:-2px 0 14px"><b>At a glance:</b> spends the most time in <b>{{ d.insight.top_stage }}</b> ({{ hdur(d.insight.top_stage_days) }} across all tickets){% if d.insight.reopened_tickets %} &middot; {{ d.insight.reopened_tickets }} ticket(s) reopened{% endif %}{% if d.insight.stuck_tickets %} &middot; {{ d.insight.stuck_tickets }} currently stuck &ge;10d{% endif %}.</p>
  {% endif %}
  <input id="histSearch" class="search" type="search" placeholder="Search tickets by key or summary…" autocomplete="off">
  <div class="controls">
    <span class="ctl-label">Lookback</span>
    {% for opt in [30,90,180,365,730] %}<a class="pill {{ 'ok' if days==opt else '' }}" href="?days={{ opt }}">{{ opt }}d</a>{% endfor %}
  </div>
  <div class="controls">
    <span class="ctl-label">Show</span>
    <button type="button" class="chipbtn active" data-filter="all">All</button>
    <button type="button" class="chipbtn" data-filter="stuck">Stuck</button>
    <button type="button" class="chipbtn" data-filter="reopened">Reopened</button>
    <button type="button" class="chipbtn" data-filter="bugs">Bugs</button>
    <button type="button" class="chipbtn" data-filter="open">Open</button>
    <button type="button" class="chipbtn" data-filter="completed">Completed</button>
    <label style="margin-left:12px;font-size:13px"><input type="checkbox" id="groupStatus"> Group by status</label>
  </div>
  <div class="toolbar">
    <a class="btn" href="/developer/{{ d.person|urlencode }}/history.csv?days={{ days }}" download>&#8595; Download full history (CSV)</a>
    <a class="btn-ghost" href="#" id="expandAll">Expand all</a>
    <a class="btn-ghost" href="#" id="collapseAll">Collapse all</a>
  </div>
  {% if legend_stages %}
  <div class="legend">
    {% for st in legend_stages %}<span><span class="sw" style="background:{{ stage_colors.get(st,'#888') }}"></span>{{ st }}</span>{% endfor %}
  </div>
  {% endif %}
  <div id="histCards">
  {% for t in d.tickets %}
  <div class="hist-card" data-order="{{ loop.index0 }}" data-type="{{ t.issue.type|lower }}" data-reopened="{{ 1 if t.reopened else 0 }}" data-stuck="{{ 1 if (t.days_in_current_stage and t.days_in_current_stage >= 10) else 0 }}" data-open="{{ 1 if t.issue.is_open else 0 }}" data-status="{{ t.issue.status }}">
    <div class="hist-head">
      <div>
        <a href="{{ t.issue.url }}" target="_blank" class="hist-key">{{ t.issue.key }}</a>
        <span>{{ t.issue.summary }}</span>
      </div>
      <div class="hist-tags">
        <span class="pill">{{ t.issue.type }}</span>
        <span class="pill">{{ t.issue.status }}</span>
        {% if t.reopened %}<span class="badge-rework" title="Sent back/reopened {{ t.reopened }} time(s)">&#8617; reopened {{ t.reopened }}&times;</span>{% endif %}
      </div>
    </div>
    <div class="hist-meta">
      Time worked (active): <b title="{{ t.active_days }} days">{{ hdur(t.active_days) }}</b> &middot;
      Total elapsed: <b title="{{ t.total_days }} days">{{ hdur(t.total_days) }}</b> &middot;
      {{ t.moves }} status change{{ '' if t.moves == 1 else 's' }}
      {% set c = agecls(t.days_in_current_stage) %}{% if c %}&middot; <span class="pill {{ c }}">stuck {{ hdur(t.days_in_current_stage) }} in {{ t.issue.status }}</span>{% endif %}
    </div>
    {% if t.stages %}
    <div class="journey">
      {% for s in t.stages %}
      <div class="seg" style="width:{{ s.pct }}%;background:{{ stage_colors.get(s.stage,'#888') }}" title="{{ s.stage }}: {{ hdur(s.days) }} ({{ s.pct }}%)">{% if s.pct >= 9 %}{{ hdur(s.days) }}{% endif %}</div>
      {% endfor %}
    </div>
    {% endif %}
    {% for s in t.stages %}{% set m = team_median.get(s.stage) %}{% if m and s.days >= m * 1.5 %}<div class="bench"><span class="pill warn" title="This ticket's time in {{ s.stage }} compared to the team median">&#9888; {{ s.stage }}: {{ hdur(s.days) }} = {{ (s.days / m)|round(1) }}&times; team median ({{ hdur(m) }})</span></div>{% endif %}{% endfor %}
    <div class="hist-grid">
      <div>
        <div class="hist-label">Time in each status</div>
        <table class="mini">
          <tr><th>Status</th><th>Time</th></tr>
          {% for s in t.per_status %}
          <tr><td>{{ s.status }}</td><td title="{{ s.days }} days">{{ hdur(s.days) }}</td></tr>
          {% else %}<tr><td colspan="2" class="muted">No recorded time.</td></tr>{% endfor %}
        </table>
      </div>
      <div>
        <details class="txns">
          <summary>Status transition history ({{ t.moves }})</summary>
          <table class="mini">
            <tr><th>When</th><th>Change</th><th>By</th></tr>
            {% for tr in t.transitions %}
            <tr><td>{{ tr.ts.strftime('%Y-%m-%d %H:%M') }}</td>
                <td>{{ tr['from'] }} &rarr; {{ tr.to }}</td>
                <td>{{ tr.author }}</td></tr>
            {% else %}<tr><td colspan="3" class="muted">No status changes recorded.</td></tr>{% endfor %}
          </table>
        </details>
      </div>
    </div>
  </div>
  {% else %}
  <p class="muted">No tickets found for {{ d.person }} in the last {{ days }} days. Widen the range above.</p>
  {% endfor %}
  </div>
</div>
<script>
(function(){
  var container=document.getElementById('histCards');
  if(!container)return;
  var cards=Array.prototype.slice.call(container.querySelectorAll('.hist-card'));
  var searchBox=document.getElementById('histSearch');
  var groupCb=document.getElementById('groupStatus');
  var activeFilter='all';
  function matchFilter(c){
    switch(activeFilter){
      case 'stuck': return c.getAttribute('data-stuck')==='1';
      case 'reopened': return c.getAttribute('data-reopened')==='1';
      case 'bugs': return c.getAttribute('data-type')==='bug';
      case 'open': return c.getAttribute('data-open')==='1';
      case 'completed': return c.getAttribute('data-open')==='0';
      default: return true;
    }
  }
  function apply(){
    var q=(searchBox&&searchBox.value||'').toLowerCase();
    container.querySelectorAll('.group-header').forEach(function(h){h.remove();});
    var visible=cards.filter(function(c){
      return matchFilter(c) && c.textContent.toLowerCase().indexOf(q)>=0;
    });
    cards.forEach(function(c){ if(visible.indexOf(c)<0) c.style.display='none'; });
    var grouped=groupCb&&groupCb.checked;
    visible.sort(function(a,b){
      if(grouped){
        var sa=a.getAttribute('data-status'), sb=b.getAttribute('data-status');
        if(sa<sb)return -1; if(sa>sb)return 1;
      }
      return parseInt(a.getAttribute('data-order'))-parseInt(b.getAttribute('data-order'));
    });
    var last=null;
    visible.forEach(function(c){
      c.style.display='';
      if(grouped){
        var st=c.getAttribute('data-status');
        if(st!==last){
          var h=document.createElement('div'); h.className='group-header'; h.textContent=st;
          container.appendChild(h); last=st;
        }
      }
      container.appendChild(c);
    });
  }
  if(searchBox)searchBox.addEventListener('input',apply);
  if(groupCb)groupCb.addEventListener('change',apply);
  document.querySelectorAll('.chipbtn').forEach(function(btn){
    btn.addEventListener('click',function(){
      document.querySelectorAll('.chipbtn').forEach(function(b){b.classList.remove('active');});
      btn.classList.add('active'); activeFilter=btn.getAttribute('data-filter'); apply();
    });
  });
  function setAll(open){document.querySelectorAll('details.txns').forEach(function(d){d.open=open;});}
  var ea=document.getElementById('expandAll'), ca=document.getElementById('collapseAll');
  if(ea)ea.addEventListener('click',function(e){e.preventDefault();setAll(true);});
  if(ca)ca.addEventListener('click',function(e){e.preventDefault();setAll(false);});
})();
</script>
"""


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def overview():
    reports = get_reports()
    devs = sorted(reports.values(), key=lambda d: (-d.throughput, -len(d.in_progress)))
    total_done = sum(d.throughput for d in devs)
    total_wip = sum(len(d.in_progress) for d in devs)
    all_cycle = [t.cycle_days for d in devs for t in d.completed if t.cycle_days is not None]
    all_oldest = [d.oldest_in_progress for d in devs if d.oldest_in_progress is not None]
    from statistics import median
    return render_template_string(
        OVERVIEW_TMPL, devs=devs, fmt=fmt,
        projects=", ".join(jc.PROJECT_KEYS), window=jc.WINDOW_DAYS,
        generated=time.strftime("%Y-%m-%d %H:%M"),
        total_done=total_done, total_wip=total_wip,
        team_cycle=fmt(round(median(all_cycle), 1) if all_cycle else None),
        team_oldest=fmt(max(all_oldest) if all_oldest else None),
    )


def _dev_filters(args):
    """Parse the developer-page filter args, shared by the page and CSV routes."""
    types = args.getlist("type")
    statuses = args.getlist("status")
    try:
        min_age = max(int(args.get("min_age") or 0), 0)
    except ValueError:
        min_age = 0
    return types, statuses, min_age


def _dev_options(d):
    """Distinct ticket types and statuses across all of a developer's tickets (for
    the filter checkboxes — taken from the unfiltered set so toggles never vanish)."""
    tickets = list(d.completed) + list(d.in_progress) + list(d.assigned)
    return (sorted({t.issue_type for t in tickets if t.issue_type}),
            sorted({t.status for t in tickets if t.status}))


def _filtered_report(d, types, statuses, min_age):
    """Return a new DeveloperReport with its lists narrowed by the filters.
    Min-open-age only applies to the open 'assigned' list, which carries age_days."""
    tset, sset = set(types), set(statuses)

    def keep(t):
        return (not tset or t.issue_type in tset) and (not sset or t.status in sset)

    assigned = [t for t in d.assigned if keep(t)]
    if min_age > 0:
        assigned = [t for t in assigned if (t.age_days or 0) >= min_age]
    return jc.DeveloperReport(
        name=d.name,
        completed=[t for t in d.completed if keep(t)],
        in_progress=[t for t in d.in_progress if keep(t)],
        assigned=assigned,
    )


def _attention(fd):
    """Summary alert counts for the 'needs attention' banner."""
    stuck = sum(1 for t in fd.in_progress if (t.days_in_status or 0) >= 10)
    reopened = sum(1 for t in (fd.in_progress + fd.completed) if getattr(t, "reopened", 0))
    blocked = sum(1 for t in fd.assigned
                  if cfg.stage_of(t.status, t.status_category) in cfg.BLOCKED_STAGES)
    oldest = max((t.age_days or 0 for t in fd.assigned), default=0)
    return {"stuck": stuck, "reopened": reopened, "blocked": blocked, "oldest": oldest}


@app.route("/developer/<name>")
def developer(name):
    reports = get_reports()
    d = reports.get(name)
    if not d:
        abort(404)
    types, statuses, min_age = _dev_filters(request.args)
    avail_types, avail_statuses = _dev_options(d)
    fd = _filtered_report(d, types, statuses, min_age)
    return render_template_string(
        DEV_TMPL, d=fd, fmt=fmt, hdur=hdur, agecls=agecls, window=jc.WINDOW_DAYS,
        alerts=_attention(fd),
        avail_types=avail_types, avail_statuses=avail_statuses,
        sel_types=set(types), sel_statuses=set(statuses), min_age=min_age)


@app.route("/developer/<name>/report.csv")
def developer_csv(name):
    reports = get_reports()
    d = reports.get(name)
    if not d:
        abort(404)
    types, statuses, min_age = _dev_filters(request.args)
    fd = _filtered_report(d, types, statuses, min_age)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([f"Developer report — {fd.name}"])
    w.writerow([f"Completed window: {jc.WINDOW_DAYS} days"])
    w.writerow(["Filters — types:", ", ".join(sorted(set(types))) or "all",
                "statuses:", ", ".join(sorted(set(statuses))) or "all",
                "min open age (days):", min_age])
    w.writerow([])
    w.writerow(["Summary", "Value"])
    w.writerow(["Open assigned", fd.open_count])
    w.writerow(["Completed", fd.throughput])
    w.writerow(["In progress", len(fd.in_progress)])
    w.writerow(["Median cycle (days)", fd.median_cycle])
    w.writerow([])
    w.writerow(["In progress"])
    w.writerow(["Key", "Summary", "Status", "Days in current status",
                "Time in progress (days)", "URL"])
    for t in fd.in_progress:
        w.writerow([t.key, t.summary, t.status, t.days_in_status, t.active_days, t.url])
    w.writerow([])
    w.writerow(["Completed"])
    w.writerow(["Key", "Summary", "Type", "Lead days", "Cycle days", "URL"])
    for t in fd.completed:
        w.writerow([t.key, t.summary, t.issue_type, t.lead_days, t.cycle_days, t.url])
    w.writerow([])
    w.writerow(["Currently assigned (open)"])
    w.writerow(["Key", "Summary", "Type", "Status", "Open age (days)", "URL"])
    for t in fd.assigned:
        w.writerow([t.key, t.summary, t.issue_type, t.status, t.age_days, t.url])
    safe = "".join(c if c.isalnum() else "_" for c in fd.name).strip("_") or "developer"
    return Response(buf.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=developer_{safe}.csv"})


@app.route("/developer/<name>/history")
def developer_history(name):
    days = int(request.args.get("days", 365))
    issues = R.load_issues(jc.fetch_working_set(days))
    d = R.employee_history(issues, name, since_days=days)
    present = [st for st in cfg.STAGE_ORDER
               if any(s["stage"] == st for t in d["tickets"] for s in t["stages"])]
    # Team-wide median time per stage, for the per-ticket outlier benchmarks.
    team_median = {row["stage"]: row["median_days"] for row in R.status_duration(issues)["rows"]}
    return render_template_string(HIST_TMPL, d=d, fmt=fmt, hdur=hdur, agecls=agecls,
                                  days=days, stage_colors=cfg.STAGE_COLORS, legend_stages=present,
                                  team_median=team_median)


@app.route("/developer/<name>/history.csv")
def developer_history_csv(name):
    days = int(request.args.get("days", 365))
    issues = R.load_issues(jc.fetch_working_set(days))
    d = R.employee_history(issues, name, since_days=days)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([f"Activity history — {d['person']}"])
    w.writerow([f"Tickets worked: {d['ticket_count']}",
                f"Total active days: {d['active_days_total']}", f"Lookback days: {days}"])
    w.writerow([])
    w.writerow(["Tickets worked on"])
    w.writerow(["Key", "Summary", "Type", "Current status", "Time worked (days)",
                "Total elapsed (days)", "Status changes", "URL"])
    for t in d["tickets"]:
        i = t["issue"]
        w.writerow([i.key, i.summary, i.type, i.status, t["active_days"],
                    t["total_days"], t["moves"], i.url])
    w.writerow([])
    w.writerow(["Time in each status"])
    w.writerow(["Key", "Status", "Days"])
    for t in d["tickets"]:
        for s in t["per_status"]:
            w.writerow([t["issue"].key, s["status"], s["days"]])
    w.writerow([])
    w.writerow(["Status transition history"])
    w.writerow(["Key", "Timestamp", "From", "To", "Author"])
    for t in d["tickets"]:
        for tr in t["transitions"]:
            w.writerow([t["issue"].key, tr["ts"].strftime("%Y-%m-%d %H:%M"),
                        tr["from"], tr["to"], tr["author"]])
    safe = "".join(c if c.isalnum() else "_" for c in d["person"]).strip("_") or "employee"
    return Response(buf.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=history_{safe}.csv"})


@app.route("/api/report.json")
def report_json():
    reports = get_reports()
    out = {}
    for name, d in reports.items():
        out[name] = {
            "throughput": d.throughput,
            "open_assigned": d.open_count,
            "avg_cycle_days": d.avg_cycle,
            "median_cycle_days": d.median_cycle,
            "in_progress": len(d.in_progress),
            "oldest_wip_days": d.oldest_in_progress,
            "completed": [{"key": t.key, "summary": t.summary, "lead_days": t.lead_days,
                           "cycle_days": t.cycle_days} for t in d.completed],
            "wip": [{"key": t.key, "summary": t.summary, "status": t.status,
                     "days_in_status": t.days_in_status} for t in d.in_progress],
            "assigned": [{"key": t.key, "summary": t.summary, "type": t.issue_type,
                          "status": t.status, "age_days": t.age_days} for t in d.assigned],
        }
    return jsonify(out)


@app.route("/report.xlsx")
def report_xlsx():
    import openpyxl
    from openpyxl.styles import Font

    reports = get_reports()
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws.append(["Developer", "Open assigned", "Completed", "Avg cycle (d)", "Median cycle (d)",
               "In progress", "Oldest WIP (d)"])
    for c in ws[1]:
        c.font = Font(bold=True)
    for d in sorted(reports.values(), key=lambda d: -d.throughput):
        ws.append([d.name, d.open_count, d.throughput, d.avg_cycle, d.median_cycle,
                   len(d.in_progress), d.oldest_in_progress])

    wc = wb.create_sheet("Completed")
    wc.append(["Developer", "Key", "Summary", "Type", "Lead (d)", "Cycle (d)"])
    for c in wc[1]:
        c.font = Font(bold=True)
    for d in reports.values():
        for t in d.completed:
            wc.append([d.name, t.key, t.summary, t.issue_type, t.lead_days, t.cycle_days])

    wp = wb.create_sheet("In Progress")
    wp.append(["Developer", "Key", "Summary", "Current status", "Days in status"])
    for c in wp[1]:
        c.font = Font(bold=True)
    for d in reports.values():
        for t in d.in_progress:
            wp.append([d.name, t.key, t.summary, t.status, t.days_in_status])

    wa = wb.create_sheet("Assigned (open)")
    wa.append(["Developer", "Key", "Summary", "Type", "Status", "Open age (d)"])
    for c in wa[1]:
        c.font = Font(bold=True)
    for d in reports.values():
        for t in d.assigned:
            wa.append([d.name, t.key, t.summary, t.issue_type, t.status, t.age_days])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=developer_report.xlsx"},
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
